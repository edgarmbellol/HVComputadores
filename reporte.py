from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def crear_hoja_vida(equipo):
    """
    Crea un reporte de hoja de vida para un equipo de cómputo.
    
    Args:
        equipo: Objeto Datospc con la información del equipo
        
    Returns:
        Workbook: Objeto Workbook de openpyxl con el reporte generado
    """
    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja de Vida"

    # Estilos
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    center_aligned = Alignment(horizontal='center', vertical='center')
    
    # Encabezado del hospital
    ws.merge_cells('A1:H1')
    ws['A1'] = "HOSPITAL GENERAL DE MEDELLÍN"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_aligned
    
    ws.merge_cells('A2:H2')
    ws['A2'] = "HOJA DE VIDA DE EQUIPO DE CÓMPUTO"
    ws['A2'].font = title_font
    ws['A2'].alignment = center_aligned
    
    # Información básica
    ws.merge_cells('A4:B4')
    ws['A4'] = "INFORMACIÓN BÁSICA"
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    
    basic_info = [
        ("Placa Serial", equipo.placaSerial),
        ("Responsable", equipo.responsable),
        ("Ubicación", equipo.Ubicacion),
        ("Nombre Usuario", equipo.nombreUsuario),
        ("Dirección MAC", equipo.direccionMac),
        ("VPN", equipo.VPN)
    ]
    
    for i, (label, value) in enumerate(basic_info, start=5):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = thin_border
        ws[f'B{i}'].border = thin_border
    
    # CPU y Procesador
    ws.merge_cells('D4:E4')
    ws['D4'] = "CPU Y PROCESADOR"
    ws['D4'].font = header_font
    ws['D4'].fill = header_fill
    
    cpu_info = [
        ("Fabricante CPU", equipo.cpuFabricante),
        ("Serial CPU", equipo.cpuSerial),
        ("Placa CPU", equipo.cpuPlaca),
        ("Nombre Procesador", equipo.procesadorNombre),
        ("Velocidad Procesador", equipo.procesadorVelo),
        ("Serial Procesador", equipo.procesadorSerial)
    ]
    
    for i, (label, value) in enumerate(cpu_info, start=5):
        ws[f'D{i}'] = label
        ws[f'E{i}'] = value
        ws[f'D{i}'].font = Font(bold=True)
        ws[f'D{i}'].border = thin_border
        ws[f'E{i}'].border = thin_border
    
    # Memoria RAM
    ws.merge_cells('G4:H4')
    ws['G4'] = "MEMORIA RAM"
    ws['G4'].font = header_font
    ws['G4'].fill = header_fill
    
    ram_info = [
        ("Marca RAM 1", equipo.ram1Marca),
        ("Capacidad RAM 1", equipo.ram1capacidad),
        ("Serial RAM 1", equipo.ram1Serial),
        ("Marca RAM 2", equipo.ram2Marca),
        ("Capacidad RAM 2", equipo.ram2Capacidad),
        ("Serial RAM 2", equipo.ram2Serial)
    ]
    
    for i, (label, value) in enumerate(ram_info, start=5):
        ws[f'G{i}'] = label
        ws[f'H{i}'] = value
        ws[f'G{i}'].font = Font(bold=True)
        ws[f'G{i}'].border = thin_border
        ws[f'H{i}'].border = thin_border
    
    # Almacenamiento
    ws.merge_cells('A12:B12')
    ws['A12'] = "ALMACENAMIENTO"
    ws['A12'].font = header_font
    ws['A12'].fill = header_fill
    
    storage_info = [
        ("Marca Disco", equipo.discoMarca),
        ("Capacidad Disco", equipo.discoCapacidad),
        ("Serial Disco", equipo.discoSerial),
        ("Unidad CD", equipo.unidadCD),
        ("Serial Unidad CD", equipo.unidadCDSerial)
    ]
    
    for i, (label, value) in enumerate(storage_info, start=13):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'A{i}'].border = thin_border
        ws[f'B{i}'].border = thin_border
    
    # Periféricos
    ws.merge_cells('D12:E12')
    ws['D12'] = "PERIFÉRICOS"
    ws['D12'].font = header_font
    ws['D12'].fill = header_fill
    
    peripherals_info = [
        ("Marca Monitor", equipo.monitorMarca),
        ("Modelo Monitor", equipo.monitorModelo),
        ("Serial Monitor", equipo.monitorSerial),
        ("Placa Monitor", equipo.monitorPlaca),
        ("Marca Teclado", equipo.tecladoMarca),
        ("Serial Teclado", equipo.tecladoSerial),
        ("Marca Mouse", equipo.mouseMarca),
        ("Serial Mouse", equipo.mouseSerial)
    ]
    
    for i, (label, value) in enumerate(peripherals_info, start=13):
        ws[f'D{i}'] = label
        ws[f'E{i}'] = value
        ws[f'D{i}'].font = Font(bold=True)
        ws[f'D{i}'].border = thin_border
        ws[f'E{i}'].border = thin_border
    
    # Observaciones
    ws.merge_cells('G12:H12')
    ws['G12'] = "OBSERVACIONES"
    ws['G12'].font = header_font
    ws['G12'].fill = header_fill
    
    ws.merge_cells('G13:H20')
    ws['G13'] = equipo.Observaciones
    ws['G13'].alignment = Alignment(wrap_text=True, vertical='top')
    ws['G13'].border = thin_border
    
    # Ajustar el ancho de las columnas
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    return wb 
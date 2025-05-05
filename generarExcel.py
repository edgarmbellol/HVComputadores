import csv
from db import *
from openpyxl import load_workbook

# Funcion para insertar datos en el excel
def datos_planilla(datos):
    plantilla = 'formato.xlsx'
    wb = load_workbook(plantilla)
    ws = wb["HOJA DE VIDA DE EQUIPOS"]
    # Insertar los datos en el excel

    # Nombre del responsable   
    ws["O10"] = datos[1]
    # Ubicacion   
    ws["D11"] = datos[2]
    # Nombre del equipo 
    ws["Y11"] = datos[3]
    # CPU Fabricante 
    ws["G15"] = datos[4]
    # CPU Serial 
    ws["Z15"] = datos[5]
    # CPU Placa
    ws["T15"] = datos[6]
    # Nombre Procesador 
    ws["G17"] = datos[7]
    # Velocidad procesador
    ws["T17"] = datos[8]
    # Procesador serial
    ws["Z17"] = datos[9]
    # Ram1 marca
    ws["G19"] = datos[10]
    # ram1 Capacidad
    ws["T19"] = datos[11]
    # ram1 Serial
    ws["Z19"] = datos[12]
    # ram2 Marca
    ws["G21"] = datos[13]
    # ram2 Capacidad
    ws["T21"] = datos[14]
    # ram2 Serial
    ws["Z21"] = datos[15]
    # Disco marca
    ws["G23"] = datos[16]
    # Disco capacidad
    ws["T23"] = datos[17]
    # Disco Serial
    ws["Z23"] = datos[18]
    # Unidad CD
    ws["G24"] = datos[19]
    # Unidad CD Serial
    ws["Z24"] = datos[20]

    # Monitor Marca
    ws["G26"] = datos[21]
    # Monitor Modelo
    ws["Z26"] = datos[22]
    # Monitor Serial
    ws["Z27"] = datos[23]
    # Monitor Placa
    ws["G27"] = datos[24]

    # Teclado marca
    ws["G29"] = datos[25]
    # Teclado Serial
    ws["Z30"] = datos[26]

    # Mouse Marca
    ws["G32"] = datos[27]
    # Mouse Serial
    ws["Z33"] = datos[28]

    # Nombre usuario
    ws["A39"] = datos[29]
    # Direccion mac
    ws["Y36"] = datos[30]

    ws = wb["Hoja Mantenimiento"]
    # Observaciones
    ws["A17"] = datos[31]
    
    nombre = datos[3]
    nombre = "Datos/" + nombre + ".xlsx"
    wb.save(nombre)


# traer datos de la base de datos
datos = traer_datos()

for dato in datos:
    nombre = datos_planilla(dato)



